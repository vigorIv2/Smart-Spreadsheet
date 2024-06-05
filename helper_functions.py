from pathlib import Path
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from typing import Any, Union


def get_sheet_from_excel(filename: Path, sheet_name: str) -> Worksheet:
    """
    Get a worksheet from an Excel file

    Args:
    filename (Path): The path to the Excel file
    sheet_name (str): The name of the worksheet

    Returns:
    Worksheet: The worksheet object
    """
    wb = load_workbook(filename)
    return wb[sheet_name]


def serialize_value(cell: Cell) -> str:
    value = cell.value
    return str(value)


def remove_none_key_value_pairs(d: dict[Any, Any]) -> dict[Any, Any]:
    """
    Remove key-value pairs where both the key and value are None

    Returns:
    dict: A new dictionary with None key-value pairs removed.
    """

    return {
        key: value for key, value in d.items() if not (key is None and value is None)
    }


def process_simple_table(ws: Worksheet) -> list[dict[str, Union[str, float, int]]]:
    """
    process_simple_table handles a simple spreadsheet which has one table starting from the top left corner
    Its first row is its header and the following rows are data records.
    Example:
    | Month    | Savings |
    | -------- | ------- |
    | January  | $250    |
    | February | $80     |
    | March    | $420    |
    """
    headers = [serialize_value(cell) for cell in ws[1]]

    records = []
    for row in ws.iter_rows(min_row=2):
        values = [serialize_value(cell) for cell in row]
        record = dict(zip(headers, values))
        records.append(remove_none_key_value_pairs(record))
    return records


def calculate_num_leading_space_per_level(row_headers: list[str]) -> int:
    for current_header, next_header in zip(row_headers, row_headers[1:]):
        current_spaces = len(current_header) - len(current_header.lstrip())
        next_spaces = len(next_header) - len(next_header.lstrip())
        if next_spaces != current_spaces:
            return next_spaces - current_spaces
    return 0


def process_hierarchical_table(ws: Worksheet) -> dict[str, Any]:
    """
    process_hierarchical_table handles a spreadsheet which has one table starting from the top left corner
    Its top left cell is empty. Its first row and first column are its headers.
    Its first column has hierarchical structural represented by the number of leading spaces.
    Some rows represents a category where the data cells are empty. Other rows represents actual data where data can be found in the data cells.
    Example:
    |                                              |30-Sep-23           |31-Oct-23           |30-Nov-23           |
    |----------------------------------------------|--------------------|--------------------|--------------------|
    |Assets                                        |                    |                    |                    |
    |   Current Assets                             |                    |                    |                    |
    |      Cash and Cash Equivalent                |                    |                    |                    |
    |         1060 TD Chequing Bank Account - #4092|587,881.66          |750,736.21          |453,234.78          |
    |         1061 TD AUD FX Currency-XXX-0283     |1,588.43            |17,457.51           |1,444.33            |
    |      Total Cash and Cash Equivalent          |$         589,470.09|$         768,193.72|$         454,679.11|
    |      1320 Prepaid Expenses                   |423,826.69          |233,127.50          |270,189.85          |
    |         1302 Prepaid License at Vid Australia|46,985.98           |68,985.98           |68,985.98           |
    |   Total Current Assets at Inc. and Australia |$      1,060,282.76 |$      1,070,307.20 |$         793,854.94|
    |Total Assets                                  |$      1,060,282.76 |$      1,070,307.20 |$         793,854.94|
    """

    def add_data(
        processed_table: dict[str, Any],
        nodes: list[str],
        col_headers: list[str],
        data_cells: tuple[Cell, ...],
    ) -> dict[str, Any]:
        current_level = processed_table
        for node in nodes[:-1]:
            if node not in current_level:
                print(
                    f"warning: can't find node {node} in processed table {current_level}. Creating a new node."
                )
                current_level[node] = {}
            current_level = current_level[node]

        current_level[nodes[-1]] = dict(
            zip(col_headers, [serialize_value(d) for d in data_cells])
        )
        return processed_table

    col_headers = [serialize_value(e) for e in ws[1][1:]]

    row_headers: list[str] = []
    for column in ws.iter_cols(min_col=1, max_col=1, values_only=False):
        row_headers = [serialize_value(cell) for cell in column[1:]]

    num_leading_space_per_level = calculate_num_leading_space_per_level(row_headers)

    if num_leading_space_per_level == 0:
        num_leading_space_per_level = 1

    processed_table: dict[str, Any] = {}
    nodes: list[str] = []

    # Process each row into the hierarchical structure
    for row in ws.iter_rows(min_row=2, values_only=False):
        level = (
            len(serialize_value(row[0])))
            - len(serialize_value(row[0])).lstrip()
         // num_leading_space_per_level
        label = serialize_value(row[0]).strip()
        data_cells = row[1:]

        nodes = nodes[:level]
        nodes.append(label)

        if any([c for c in data_cells if c.value is not None]):
            processed_table = add_data(processed_table, nodes, col_headers, data_cells)

    return remove_none_key_value_pairs(processed_table)
